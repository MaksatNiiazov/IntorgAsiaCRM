import calendar
import os
import time
from urllib.parse import quote
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from crm_app.forms import ServiceForm, CashboxForm, AddServiceForm, CashboxOperationForm, ServiceTypeForm, \
    ConsumablesForm, AddServiceEmployerForm, CashboxOperationCategoryForm, UserForm, DiscountForm
from crm_app.models import Order, Service, Cashbox, OrderService, CustomUser, CashboxOperation, CashboxCategory, \
    ServiceType, ServiceOrder, EmployerOrder, Consumables, ModelChangeLog, OrderConsumables
from datetime import date, timedelta
from django.db.models import Sum, Count, Q

from crm_warehouse.models import ProductService, EmployerProduct, Product
from users.models import DiscountType
from users.permissions import WorkerRequiredMixin

class Locked(LoginRequiredMixin):
    login_url = "login"


class LockedView(LoginRequiredMixin, WorkerRequiredMixin):
    login_url = "login"




class StatisticView(LockedView, ListView):
    model = Order
    template_name = 'crmapp/сhart.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        filter_type = self.request.GET.get('filter_type')
        filter_start = self.request.GET.get('start_date')
        filter_end = self.request.GET.get('end_date') or date.today()

        if filter_type == 'week':
            if not filter_start or not filter_end:
                today = date.today()
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                self.request.GET = self.request.GET.copy()
                self.request.GET['start_date'] = start_date.strftime('%Y-%m-%d')
                self.request.GET['end_date'] = end_date.strftime('%Y-%m-%d')
        elif filter_type == 'month':
            if not filter_start:
                today = date.today()
                start_date = date(today.year, today.month, 1)
                end_date = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
                self.request.GET = self.request.GET.copy()
                self.request.GET['start_date'] = start_date.strftime('%Y-%m-%d')
                self.request.GET['end_date'] = end_date.strftime('%Y-%m-%d')
        elif filter_type == 'quarter':
            if not filter_start:
                today = date.today()
                quarter_start_month = ((today.month - 1) // 3) * 3 + 1
                start_date = date(today.year, quarter_start_month, 1)
                end_date = date(today.year, quarter_start_month + 2,
                                calendar.monthrange(today.year, quarter_start_month + 2)[1])
                self.request.GET = self.request.GET.copy()
                self.request.GET['start_date'] = start_date.strftime('%Y-%m-%d')
                self.request.GET['end_date'] = end_date.strftime('%Y-%m-%d')
        elif filter_type == 'year':
            if not filter_start:
                today = date.today()
                start_date = date(today.year, 1, 1)
                end_date = date(today.year, 12, 31)
                self.request.GET = self.request.GET.copy()
                self.request.GET['start_date'] = start_date.strftime('%Y-%m-%d')
                self.request.GET['end_date'] = end_date.strftime('%Y-%m-%d')

        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date') or date.today()

        queryset = queryset.filter(date__range=[start_date, end_date]).values('year', 'month', 'day').annotate(
            total_count=Count('id'), total_amount=Sum('amount'), total_cost_price=Sum('cost_price'), profit=(
                    Sum('amount') - Sum('cost_price'))).distinct()

        return queryset

    def get_context_data(self, **kwargs):
        queryset = self.get_queryset()
        total_amount = queryset.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        total_cost_price = queryset.aggregate(total_cost_price=Sum('total_cost_price'))['total_cost_price'] or 0
        total_count = queryset.aggregate(total_count=Count('id'))['total_count'] or 0
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date') or ""
        end_date = self.request.GET.get('end_date') or date.today().strftime('%Y-%m-%d')
        filter_type = self.request.GET.get('filter_type')
        context['start_date'] = start_date
        context['end_date'] = end_date
        context['filter_type'] = filter_type
        context['total_amount'] = total_amount
        context['total_cost_price'] = total_cost_price
        context['total_count'] = total_count
        context['profit'] = total_amount - total_cost_price or 0

        return context


class OrderListView(LockedView, ListView):
    model = Order
    template_name = 'crmapp/order_list.html'
    ordering = ['-date']
    paginate_by = 20

    def get_queryset(self):
        return Order.objects.exclude(stage='closed')


class ClosedOrderListView(LockedView, ListView):
    model = Order
    template_name = 'crmapp/order_list.html'
    ordering = ['-date']
    paginate_by = 20

    def get_queryset(self):
        return Order.objects.filter(stage='closed')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(ClosedOrderListView, self).get_context_data()
        context['closed'] = True
        return context


class OrderDetailView(LockedView, DetailView):
    model = Order
    template_name = 'crmapp/order_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['services'] = ServiceOrder.objects.filter(order=self.object.id)
        context['products'] = Product.objects.filter(order=self.object.id)
        context['consumables_in_order'] = OrderConsumables.objects.filter(order=self.object.id)
        context['cashboxes'] = Cashbox.objects.all()
        if self.object.amount == 0 and self.object.stage == 'dispatched':
            context['next_stage'] = True
        order = self.get_object()
        revenue = order.amount - order.cost_price

        context['revenue'] = revenue
        return context


class ArrivalShipmentUpdate(View):
    def post(self, request, pk):
        order = Order.objects.get(id=pk)
        date_of_actual_arrival = self.request.POST.get('date_of_actual_arrival') or None
        actual_shipment_date = self.request.POST.get('actual_shipment_date') or None
        order.date_of_actual_arrival = date_of_actual_arrival if date_of_actual_arrival is not None else order.date_of_actual_arrival
        order.actual_shipment_date = actual_shipment_date if actual_shipment_date is not None else order.actual_shipment_date
        order.save()
        print(order.actual_shipment_date, order.date_of_actual_arrival)
        print(self.request.META.get('HTTP_REFERER'))
        return redirect(self.request.META.get('HTTP_REFERER'))


class SertviceToOrderView(LockedView, View):

    def get(self, request):
        context = {
            'services': Service.objects.filter(acceptance=False, single=False),
            'employers': CustomUser.objects.filter(user_type='worker')
        }
        return render(request, 'crmapp/service_add.html', context)

    def post(self, request):
        pass


class AddServiceView(LockedView, CreateView):
    form_class = AddServiceForm
    template_name = 'crmapp/service_add.html'

    def form_valid(self, form):
        order = form.cleaned_data['order']
        service = form.cleaned_data['service']
        count = form.cleaned_data['count']
        update_order = Order.objects.get(id=order.pk)
        service_order, _ = ServiceOrder.objects.get_or_create(order=order, service=service)
        client = update_order.client

        old_count = service_order.count or 0
        old_amount = old_count * service.price or 0
        old_cost = old_count * service.cost_price or 0
        new_count = count
        new_amount = count * service.price
        new_cost = count * service.cost_price

        count_diff = new_count - old_count
        amount_diff = new_amount - old_amount
        cost_diff = new_cost - old_cost

        update_order.amount += amount_diff
        update_order.cost_price += cost_diff

        service_order.count += count_diff
        service_order.price += amount_diff

        client.money += amount_diff
        client.profit += amount_diff - cost_diff

        update_order.save(update_fields=['count', 'amount', 'cost_price'])
        service_order.save(update_fields=['count', 'price'])
        client.save(update_fields=['money', 'profit'])
        return redirect('invoice_generation', pk=order.pk)

    def form_invalid(self, form):
        order = form.cleaned_data['order']
        print(form.errors)
        return redirect('invoice_generation', pk=order.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_id'] = self.kwargs['order_id']
        context['services'] = Service.objects.all()
        context['employers'] = CustomUser.objects.filter(user_type='worker')
        return context


class AddServiceEmployerView(LockedView, CreateView):

    form_class = AddServiceEmployerForm
    template_name = 'crmapp/service_add.html'

    def form_valid(self, form):
        product = form.cleaned_data['product']
        order = product.order
        client = order.client
        count = form.cleaned_data['service_count']
        service = Service.objects.get(id=int(self.request.POST.get('service')))
        employer = form.cleaned_data['employer']
        price = service.price * count
        cost_price = service.cost_price * count

        employer_product, _ = EmployerProduct.objects.get_or_create(employer=employer, product=product,)
        employer_product.service_count = count
        employer_product.save()

        product_service = ProductService.objects.create(service=service, employer_product=employer_product,
                                                        count=count,)
        employer_order, _ = EmployerOrder.objects.get_or_create(order=order, user=employer)
        service_order, _ = ServiceOrder.objects.get_or_create(order=order, service=service)
        service_order.count += count
        service_order.price += price
        service_order.cost_price += cost_price
        service_order.save()

        order.amount += price
        order.cost_price += cost_price
        order.count += count
        order.save()

        client.money += price
        client.profit += price - cost_price
        client.save()

        employer.money += cost_price
        employer.save()

        employer_order.service_count += count
        employer_order.salary += price

        employer_order.save()

        return redirect('quality_check', pk=order.pk)

    def form_invalid(self, form):

        print(form.errors)
        return redirect('invoice_generation', pk=form.product.order.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_id'] = self.kwargs['order_id']
        context['services'] = Service.objects.all()
        context['employers'] = CustomUser.objects.filter(user_type='worker')
        return context


class OrderServiceDeleteView(LockedView, DeleteView):
    model = OrderService

    def form_valid(self, form):
        order = Order.objects.get(pk=self.object.order.id)
        order.amount -= self.object.salary
        order.count -= self.object.count
        order.cost_price -= self.object.service.cost_price
        order.save()
        client = CustomUser.objects.get(id=order.client_id)
        client.money -= self.object.salary
        client.product_count -= self.object.count
        client.save()
        worker = CustomUser.objects.get(id=self.object.employer_id)
        worker.money -= self.object.salary
        worker.product_count -= self.object.count
        worker.save()
        cashbox = Cashbox.objects.get(id=order.cashbox_id)
        cashbox.balance -= self.object.salary
        cashbox.save()
        self.object.delete()
        # ModelChangeLog.add_log(model_name='сервис', user_id=self.request.user.id, old_value=f'{service.name}/{old_count}', new_value=f'{service.name}/{new_count}')

        return redirect('order_detail', pk=self.object.order.pk)


class ServiceListView(LockedView, ListView):
    model = Service
    template_name = 'crmapp/service_list.html'
    paginate_by = 25

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(ServiceListView, self).get_context_data()
        context['types'] = ServiceType.objects.all()
        return context


class ServiceTypeListView(LockedView, ListView):
    model = ServiceType
    template_name = 'crmapp/service_types.html'
    success_url = '/service/type/list/'


class ServiceTypeCreateView(LockedView, CreateView):
    model = ServiceType
    form_class = ServiceTypeForm
    template_name = 'crmapp/service_types.html'
    success_url = '/service/type/list/'


class ServiceTypeUpdateView(LockedView, UpdateView):
    model = ServiceType
    form_class = ServiceTypeForm
    template_name = 'crmapp/service_types.html'
    success_url = '/service/type/list/'


class ServiceTypeDeleteView(LockedView, DeleteView):
    model = ServiceType
    template_name = 'crmapp/service_types.html'
    success_url = '/service/type/list/'


class ServiceCreateView(LockedView, CreateView):
    model = Service
    form_class = ServiceForm
    template_name = 'crmapp/service_list.html'
    success_url = '/services/'

    def form_valid(self, form):
        response = super().form_valid(form)
        object = self.object
        object.before_defective = (self.request.POST.get('gender'))
        object.save()
        return response


class ServiceUpdateView(LockedView, UpdateView):
    model = Service
    form_class = ServiceForm
    template_name = 'crmapp/service_list.html'
    success_url = '/services/'

    def form_valid(self, form):
        object = self.object
        object.before_defective = (self.request.POST.get('gender'))
        object.save()
        response = super().form_valid(form)
        return response


class ServiceDeleteView(LockedView, DeleteView):
    model = Service
    success_url = '/services/'


class ConsumablesListView(LockedView, ListView):
    model = Consumables
    template_name = 'crmapp/consumable_list.html'


class ConsumablesCreateView(LockedView, CreateView):
    model = Consumables
    form_class = ConsumablesForm
    success_url = '/consumables/'


class ConsumablesUpdateView(LockedView, UpdateView):
    model = Consumables
    form_class = ConsumablesForm
    template_name = 'crmapp/consumable_list.html'
    success_url = '/consumables/'


class ConsumablesDeleteView(LockedView, DeleteView):
    model = Consumables
    success_url = '/consumables/'


class EmployerListView(LockedView, ListView):
    model = CustomUser
    template_name = 'crmapp/employer_list.html'
    paginate_by = 20
    ordering = 'count'

    def get_queryset(self):
        return CustomUser.objects.filter(user_type='worker')


class EmployerDetailView(LockedView, DetailView):
    model = CustomUser
    template_name = 'crmapp/employer_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['orders'] = EmployerOrder.objects.filter(user_id=self.object.id)
        context['cashboxes'] = Cashbox.objects.all()
        return context


class PayASalaryView(LockedView, View):

    def post(self, request, pk):
        emp_order = EmployerOrder.objects.get(id=pk)
        user = CustomUser.objects.get(id=emp_order.user.id)
        num = int(self.request.POST.get('num'))
        cashbox = Cashbox.objects.get(id=int(self.request.POST.get('cashbox')))
        category = CashboxCategory.objects.get_or_create(category="Зарплата")[0]
        balance_check = cashbox.balance - num
        client = emp_order.order.client


        if balance_check < 0:
            messages.error(self.request, "В кассе недостаточно средств!")
            return redirect('employer_detail', emp_order.user.id)
        CashboxOperation.objects.create(user_id=self.request.user.id, category=category, cashbox_from=cashbox,
                                        money=num, comment=f'зарплата {user}')
        cashbox.balance -= num
        user.money -= num
        emp_order.salary -= num
        cashbox.save()
        user.save()
        emp_order.save()

        ModelChangeLog.add_log(model_name='зарплата', user_id=self.request.user.id, change_type='зарплата',
                               old_value=f'{user}', new_value=f'{num}')

        return redirect('employer_detail', emp_order.user.id)


class EmployerOrderView(LockedView, ListView):
    model = EmployerProduct
    template_name = 'crmapp/employer_order.html'

    def get_queryset(self):
        order = Order.objects.get(id=self.kwargs['order_id'])
        query_set = EmployerProduct.objects.filter(product__order=order, employer_id=self.kwargs['pk'])

        return query_set

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(EmployerOrderView, self).get_context_data()
        context['order_id'] = self.kwargs['order_id']
        context['user'] = CustomUser.objects.get(id=self.kwargs['pk'])
        return context


class ClientListView(LockedView, ListView):
    model = CustomUser
    template_name = 'crmapp/client_list.html'
    paginate_by = 20

    def get_queryset(self):
        return CustomUser.objects.filter(user_type='client')


class ClientDetailView(LockedView, DetailView):
    model = CustomUser
    template_name = 'crmapp/client_detail.html'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super(ClientDetailView, self).get_context_data()
        context['orders'] = Order.objects.filter(client_id=self.object.id)
        return context


class UserChangeView(LockedView, UpdateView):
    template_name = 'crmapp/user_change.html'
    model = CustomUser
    form_class = UserForm

    def get_success_url(self):
        referer = self.request.META.get('HTTP_REFERER')
        if referer:
            return referer
        else:
            return reverse('/')


class DiscountListView(LockedView, ListView):
    model = DiscountType
    template_name = 'crmapp/discount.html'


class DiscountCreateView(LockedView, CreateView):
    model = DiscountType
    form_class = DiscountForm
    template_name = 'crmapp/discount.html'
    success_url = '/discounts/'


class DiscountUpdateView(LockedView, UpdateView):
    model = DiscountType
    form_class = DiscountForm
    template_name = 'crmapp/discount.html'

    def get_success_url(self):
        referer = self.request.META.get('HTTP_REFERER')
        if referer:
            return referer
        else:
            return reverse('/')


class DiscountDeleteView(LockedView, DeleteView):
    model = DiscountType
    template_name = 'crmapp/discount.html'
    success_url = '/discounts/'


class CashboxListView(LockedView, ListView):
    model = Cashbox
    template_name = 'crmapp/cashbox_list.html'
    paginate_by = 20


class CashboxDetailView(LockedView, DetailView):
    model = Cashbox
    template_name = 'crmapp/cashbox_detail.html'

    def get_context_data(self, **kwargs):
        context = super(CashboxDetailView, self).get_context_data()
        operations = CashboxOperation.objects.all()
        context['operations_from'] = operations.filter(cashbox_from=self.object).order_by('-id')[:10]
        context['operations_to'] = operations.filter(cashbox_to=self.object).order_by('-id')[:10]
        context['users'] = CustomUser.objects.filter(user_type='worker')
        cashboxes = Cashbox.objects.all()
        context['cashboxes'] = cashboxes.exclude(id=self.object.id)
        context['categories'] = CashboxCategory.objects.all()
        return context


class CashboxCreateView(LockedView, CreateView):
    model = Cashbox
    form_class = CashboxForm
    template_name = 'crmapp/cashbox_list.html'
    success_url = '/cashboxes/'

    def form_valid(self, form):
        response = super().form_valid(form)
        model_name = 'касса'
        user = self.request.user.id
        change_type = 'создание'
        new_value = form.instance.name
        ModelChangeLog.add_log(model_name, user, change_type=change_type, new_value=new_value)

        return response


class CashboxUpdateView(LockedView, UpdateView):
    model = Cashbox
    form_class = CashboxForm
    template_name = 'crmapp/cashbox_list.html'
    success_url = '/cashboxes/'


class CashboxDeleteView(LockedView, DeleteView):
    model = Cashbox
    success_url = '/cashboxes/'


class OperationCategoryListView(LockedView, ListView):
    model = CashboxCategory
    template_name = 'crmapp/cashbox_operation_types.html'
    paginate_by = 20


class OperationCategoryCreateView(LockedView, CreateView):
    model = CashboxCategory
    form_class = CashboxOperationCategoryForm
    template_name = 'crmapp/cashbox_operation_types.html'
    success_url = '/operation/categories/'

    def form_valid(self, form):
        response = super().form_valid(form)
        model_name = 'кстегории операций+'
        user = self.request.user.id
        change_type = 'создание'
        new_value = form.instance.category
        ModelChangeLog.add_log(model_name, user, change_type=change_type, new_value=new_value)

        return response


class OperationCategoryViewUpdateView(LockedView, UpdateView):
    model = CashboxCategory
    form_class = CashboxOperationCategoryForm
    success_url = '/operation/categories/'


class OperationCategoryDeleteView(LockedView, DeleteView):
    model = CashboxCategory
    success_url = '/operation/categories/'
    template_name = 'crmapp/cashbox_operation_types.html'


class CashBoxAddOperationView(LockedView, CreateView):
    model = CashboxOperation
    form_class = CashboxOperationForm
    template_name = 'crmapp/cashbox_detail.html'

    def form_valid(self, form):
        cashbox_to = form.cleaned_data['cashbox_to']
        if not cashbox_to:
            cashbox_to = None
        cashbox_from = form.cleaned_data['cashbox_from']
        if not cashbox_from:
            cashbox_from = None
        category = form.cleaned_data['category']
        if self.request.POST.get('check') == 'to':
            money = form.cleaned_data['money']
            if money <= 0:
                messages.error(self.request, "Вы пытаетесь сделать приход на 0!")
                return redirect(self.request.META.get('HTTP_REFERER'))
            comment = form.cleaned_data['comment']
            if not comment:
                comment = ''
            operation = CashboxOperation.objects.create(
                user_id=self.request.user.id,
                category=category,
                money=form.cleaned_data['money'],
                comment=comment,
                cashbox_from=cashbox_from,
                cashbox_to=cashbox_to,
            )
            cashbox = Cashbox.objects.get(id=cashbox_to.id)
            old_v = cashbox.balance
            cashbox.balance += money
            cashbox.save()
            ModelChangeLog.add_log(model_name=f'касса {cashbox.name}', user_id=self.request.user.id,
                                   change_type=f'приход({money})', old_value=f'{old_v}', new_value=f'{cashbox.balance}')

        elif self.request.POST.get('check') == 'from':
            money = form.cleaned_data['money']
            if money <= 0:
                messages.error(self.request, "Вы пытаетесь сделать расход на 0!")
                return redirect(self.request.META.get('HTTP_REFERER'))
            comment = form.cleaned_data['comment']
            if not comment:
                comment = ''

            cashbox = Cashbox.objects.get(id=cashbox_from.id)
            result = cashbox.balance - money
            if result < 0:
                messages.error(self.request, "В кассе недостаточно средств!")
                return redirect(self.request.META.get('HTTP_REFERER'))
            old_v = cashbox.balance
            cashbox.balance = result
            operation = CashboxOperation.objects.create(
                user_id=self.request.user.id,
                category=form.cleaned_data['category'],
                money=money,
                comment=comment,
                cashbox_from=cashbox_from,
                cashbox_to=cashbox_to,
            )
            cashbox.save()
            ModelChangeLog.add_log(model_name=f'касса {cashbox.name}', user_id=self.request.user.id,
                                   change_type=f'расход({money})', old_value=f'{old_v}', new_value=f'{result}')

        return redirect(self.request.META.get('HTTP_REFERER'))

    def form_invalid(self, form):
        print(form.errors)
        return redirect(self.request.META.get('HTTP_REFERER'))


class CashboxOperationFromListView(LockedView, ListView):
    model = CashboxOperation
    template_name = 'crmapp/cashbox_operation_list.html'
    paginate_by = 50

    def get_queryset(self):
        return CashboxOperation.objects.filter(cashbox_from_id=self.kwargs['pk'])


class CashboxOperationToListView(CashboxOperationFromListView):

    def get_queryset(self):
        return CashboxOperation.objects.filter(cashbox_to_id=self.kwargs['pk'])


class CashboxExport(LockedView, View):
    def post(self, request):
        start_date = self.request.POST['start_date']
        end_date = self.request.POST['end_date']
        cashbox = self.request.POST['cashcox']
        cashbox_operations = CashboxOperation.objects.filter(
            Q(cashbox_from=cashbox) | Q(cashbox_to=cashbox),
            date__range=(start_date, end_date)
        )
        file_name = 'cashbox_operations_template.xlsx'  # Excel template file name
        root_directory = 'excel'  # Root directory to search for the template file
        file_path = None

        for root, dirs, files in os.walk(root_directory):
            if file_name in files:
                file_path = os.path.join(root, file_name)
                break

        if not file_path:
            return HttpResponse('File not found')

        wb = load_workbook(file_path)
        sheet = wb.active

        # Set column headers
        sheet['A1'] = 'Дата'
        sheet['B1'] = 'Время'
        sheet['C1'] = 'От'
        sheet['D1'] = 'Сумма'
        sheet['E1'] = 'Куда'
        sheet['F1'] = 'Категория'
        sheet['G1'] = 'Сотрудник'
        sheet['H1'] = 'Комментарий'

        # Set column widths and formatting
        column_widths = [20, 20, 10, 30, 20, 20, 15, 15]
        for col_num, width in enumerate(column_widths, start=1):
            col_letter = chr(ord('A') + col_num - 1)
            sheet.column_dimensions[col_letter].width = width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center')
        font = Font(color="000000")

        row = 2
        for operation in cashbox_operations:
            # Set cell values
            sheet[f'A{row}'] = operation.date.strftime('%Y-%m-%d')
            sheet[f'B{row}'] = operation.time.strftime('%H:%M')
            sheet[f'C{row}'] = operation.cashbox_from.name if operation.cashbox_from else ''
            sheet[f'D{row}'] = operation.money
            sheet[f'E{row}'] = operation.cashbox_to.name if operation.cashbox_to else ''
            sheet[f'F{row}'] = operation.category.category if operation.category else ''
            sheet[f'G{row}'] = operation.user.username if operation.user else ''
            sheet[f'H{row}'] = operation.comment if operation.comment else ''

            # Apply formatting to cells
            for col_num in range(1, 9):
                cell = sheet.cell(row=row, column=col_num)
                cell.border = thin_border
                cell.alignment = alignment

            row += 1

        # Save the workbook
        timestamp = int(time.time())  # Add a timestamp to the file name to make it unique
        new_file_path = f'excel/cashbox_operations_{timestamp}.xlsx'
        wb.save(new_file_path)
        wb.close()

        # Prepare the response for file download
        with open(new_file_path, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f'CashboxOperations_{date.today()}.xlsx'
            quoted_filename = quote(filename, encoding='utf-8')
            response['Content-Disposition'] = f'attachment; filename="{quoted_filename}"'

        # Delete the newly generated file
        max_retries = 3
        retry_delay = 1  # Delay in seconds
        retry_count = 0

        while retry_count < max_retries:
            try:
                os.remove(new_file_path)
                break
            except PermissionError:
                retry_count += 1
                time.sleep(retry_delay)
        ModelChangeLog.add_log(model_name=f'касса {cashbox}', user_id=self.request.user.id,
                               change_type='экспорт')
        return response


class MakeAPaymentView(LockedView, View):
    def post(self, request):
        order = Order.objects.get(id=self.request.POST.get('order'))
        cashbox = Cashbox.objects.get(id=self.request.POST.get('cashbox'))
        money = int(self.request.POST.get('money'))

        check = order.amount - money
        if check < 0:
            messages.error(self.request, "Вы пытаетесь провести оплату на сумму бользую стоимости заказа!")
            return redirect(self.request.META.get('HTTP_REFERER'))
        if money <= 0:
            messages.error(self.request, "Вы пытаетесь провести оплату на 0!")
            return redirect(self.request.META.get('HTTP_REFERER'))
        operation_category, _ = CashboxCategory.objects.get_or_create(category='оплата заказа')
        operation = CashboxOperation.objects.create(
            user_id=self.request.user.id,
            category=operation_category,
            money=money,
            comment=f'оплата заказа №{order.id}',
            cashbox_to=cashbox,
        )
        old_v = cashbox.balance
        cashbox.balance += money
        cashbox.save()

        order.amount -= money
        order.amount_paid += money
        order.save()
        client = order.client
        if client.referral:
            referal = client.referral
            percent = 10
            services = ServiceOrder.objects.filter(order=order)
            new_amount = 0
            for sirvice in services:
                if sirvice.service.discount:
                    new_amount += (sirvice.price / 100) * percent
                else:
                    new_amount += sirvice.price
                referal.referal_money = round(new_amount)
                order.referral_money = new_amount
                order.save()
                referal.save()
        ModelChangeLog.add_log(model_name=f'касса {cashbox.name}', user_id=self.request.user.id,
                               change_type=f'оплата заказа №{order.id} ({money})', old_value=f'{old_v}',
                               new_value=f'{cashbox.balance}')

        return redirect('order_detail', order.id)


class ShowLogsView(LockedView, ListView):
    model = ModelChangeLog
    template_name = 'crmapp/logs.html'
    paginate_by = 50
    ordering = '-id'
