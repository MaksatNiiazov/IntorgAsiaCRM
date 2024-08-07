import math
import os
import time
import requests
from datetime import date
from itertools import groupby
from urllib.parse import quote
import pandas as pd
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views import View
from django.views.generic import CreateView, UpdateView, DetailView, ListView
from django.contrib import messages
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from crm_app.views import LockedView, Locked
from crm_warehouse.forms import UploadForm, AcceptanceForm, ProductForm, ProductUnpackingForm, EmployerProductForm, \
    DefectiveCheckForm, BarcodeForm
from crm_warehouse.models import Product, EmployerProduct, SetOfServices, ServiceInSet, ProductService
from crm_app.models import Order, OrderStages, Service, OrderService, ServiceOrder, EmployerOrder, Consumables, \
    OrderConsumables
from users.models import User as CustomUser


class DashboardView(Locked, ListView):
    model = Order
    template_name = 'stages/dashboard.html'
    context_object_name = 'orders'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stages = OrderStages.choices
        context['stages'] = stages
        return context

    def get_queryset(self):
        if self.request.user.user_type == 'client':
            queryset = Order.objects.filter(client_id=self.request.user.id)
        else:
            queryset = super().get_queryset()
        queryset = queryset.annotate(total_declared_quantity=Sum('products__declared_quantity'))
        queryset = queryset.annotate(total_actual_quantity=Sum('products__actual_quantity'))

        return queryset


class AcceptanceView(LockedView, View):
    template_name = 'stages/acceptance.html'

    def get(self, request):
        users = CustomUser.objects.filter(user_type='client')
        context = {
            'users': users,
            'referrals': CustomUser.objects.all(),

        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = AcceptanceForm(request.POST)
        if form.is_valid():
            # Обработка данных из формы POST
            client = form.cleaned_data['client']
            name = form.cleaned_data['name']
            count = form.cleaned_data['count']
            order = Order.objects.create(name=name, client=client, count=count)
            order.month = date.today().month
            order.year = date.today().year
            order.day = date.today().day
            if order.discount:
                order.discount = client.discount.percent
            order.discount = 0
            order.save()
            return redirect('dashboard')

        else:
            print(form.errors)

            return redirect('dashboard')


class ImportExcelView(View):
    template_name = 'stages/database_loading.html'

    def get(self, request, order_id):
        form = UploadForm()
        products = Product.objects.filter(order_id=order_id)
        context = {
            'order_id': order_id,
            'products': products,
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request, order_id):
        form = UploadForm(request.POST, request.FILES)
        products = Product.objects.filter(order_id=order_id)
        context = {
            'order_id': order_id,
            'products': products,
            'form': form
        }
        if form.is_valid():
            excel_file = form.cleaned_data['file']

            try:
                file_path = 'crm_warehouse/excel/'
                file_full_path = os.path.join(file_path, excel_file.name)
                with open(file_full_path, 'wb') as file:
                    for chunk in excel_file.chunks():
                        file.write(chunk)
                df = pd.read_excel(file_full_path, skiprows=1)

                # Fill NaN values with previous non-null values per column
                columns_to_ffill = ['Цвет ', 'Название товара', 'Страна', 'Состав (Материал)', 'Бренд',
                                    'Комментарий (Тех задание)']
                df[columns_to_ffill] = df[columns_to_ffill].ffill()
                products_to_create = []

                for _, row in df.iterrows():
                    barcode = row.get('Баркод (Штрихкод)', None)
                    if barcode and not math.isnan(barcode):
                        count = row.get('Факт', 0)
                        if isinstance(count, (float, int)) and math.isnan(count):
                            count = 0
                        declared_quantity = row.get('Заявленное количество', 0)
                        if isinstance(declared_quantity, (float, int)) and math.isnan(declared_quantity):
                            declared_quantity = 0

                        size = row['Размер ']
                        color = row['Цвет ']
                        composition = row['Состав (Материал)']
                        brand = row['Бренд']
                        country = row['Страна']
                        comment = row['Комментарий (Тех задание)']

                        size = '' if pd.isna(size) else size
                        color = '' if pd.isna(color) else color
                        composition = '' if pd.isna(composition) else composition
                        brand = '' if pd.isna(brand) else brand
                        country = '' if pd.isna(country) else country
                        comment = '' if pd.isna(comment) else comment

                        product = Product(
                            barcode=barcode,
                            article=row['Ваш артикул на ВБ (Номенклатура)'],
                            order_id=order_id,
                            name=row['Название товара'],
                            count=count,
                            declared_quantity=declared_quantity,
                            actual_quantity=declared_quantity,
                            size=size,
                            color=color,
                            composition=composition,
                            brand=brand,
                            defective=0,
                            good_quality=0,
                            country=country,
                            comment=comment,

                            confirmation=False,
                            in_work=False
                        )
                        products_to_create.append(product)
                Product.objects.bulk_create(products_to_create)

                os.remove(file_full_path)

                return redirect(reverse('import_excel', args=[order_id]))

            except Exception as e:
                error_message = f"Произошла ошибка при обработке файла Excel: {str(e)}"
                context['error_message'] = error_message

        return render(request, self.template_name, context)


class ProductDeleteView(LockedView, View):

    def post(self, request, pk):
        product = Product.objects.get(id=pk)
        order = product.order.id
        product.delete()
        return redirect('import_excel', order)


class ProductAddView(LockedView, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'stages/database_loading.html'

    def form_valid(self, form):
        product = Product(
            barcode=form.cleaned_data['barcode'],
            article=form.cleaned_data['article'],
            order=form.cleaned_data['order'] or 1,
            name=form.cleaned_data['name'],
            count=form.cleaned_data['count'] or 0,
            declared_quantity=form.cleaned_data['declared_quantity'] or 0,
            size=form.cleaned_data['size'] or '',
            color=form.cleaned_data['color'] or '',
            composition=form.cleaned_data['composition'] or '',
            brand=form.cleaned_data['brand'] or '',
            country=form.cleaned_data['country'] or '',
            confirmation=False,
            in_work=False
        )

        product.save()
        return redirect('import_excel', product.order.id)

    def form_invalid(self, form):
        print(form.errors)
        return self.success_url

    def get_success_url(self):
        product_id = self.kwargs['order_id']
        return reverse('product_detail', args=[product_id])


class UnpackingView(LockedView, DetailView):
    template_name = 'stages/unpacking.html'
    model = Order

    def get_context_data(self, **kwargs):
        context = super(UnpackingView, self).get_context_data()
        context['products'] = Product.objects.filter(order_id=self.object.id)

        return context


class ProductUpdateView(LockedView, UpdateView):
    model = Product
    form_class = ProductUnpackingForm
    template_name = 'stages/unpacking.html'

    def get_success_url(self):
        return reverse('unpacking', self.object.order.id)

    def form_valid(self, form):
        product = self.object
        product.actual_quantity = form.cleaned_data['actual_quantity']
        product.good_quality = form.cleaned_data['actual_quantity']
        product.comment = form.cleaned_data['comment']
        product.confirmation = form.cleaned_data['confirmation']
        product.save()

        return redirect('unpacking', self.object.order.id)

    def form_invalid(self, form):
        print(form.errors)
        return redirect('unpacking', self.object.order.id)


class QualityCheckView(LockedView, DetailView):
    template_name = 'stages/quality_check.html'
    model = Order

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['products_in_work'] = Product.objects.filter(order=self.object, in_work=True)
        context['products'] = Product.objects.filter(order=self.object, in_work=False)
        context['employers'] = CustomUser.objects.filter(user_type='worker')
        context['services'] = Service.objects.filter(acceptance=False, single=False)
        context['order_id'] = self.object.id
        context['sets_of_services'] = SetOfServices.objects.filter(order_id=self.object.id)

        return context


class QualityUpdateView(LockedView, CreateView):
    model = Order
    form_class = EmployerProductForm
    template_name = 'stages/quality_check.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['products'] = Product.objects.filter(order=self.object, in_work=False)
        context['products_in_work'] = Product.objects.filter(order=self.object, in_work=True)
        context['workers'] = CustomUser.objects.filter(user_type='worker')
        context['order_id'] = self.object.order.id
        return context

    def form_valid(self, form):
        product_id = self.request.POST.get('product')
        product = Product.objects.get(id=product_id)
        product.in_work = True
        product.save()
        order = product.order
        employer = form.cleaned_data['employer']
        employer_product = EmployerProduct.objects.get_or_create(
            product=product,
            employer=employer,
        )[0]
        employer_product.product_count += product.good_quality
        employer_product.save()

        return redirect('quality_check', order.id)

    def form_invalid(self, form):
        print(form.errors)
        return redirect('quality_check', self.object.id)


class SetOfServiceCreateView(LockedView, View):
    def get(self, request, pk):
        services = Service.objects.filter(single=False, acceptance=False)
        services_before = services.filter(before_defective=True)
        services_after = services.filter(before_defective=False)
        sets = SetOfServices.objects.filter(order_id=pk)

        context = {
            'order_id': pk,
            'services_before': services_before,
            'services_after': services_after,
            'sets': sets,
        }
        return render(request, 'stages/sets_of_services/set_of_srvices.html', context)

    def post(self, request, pk):
        name = self.request.POST.get('name')
        services = self.request.POST.getlist('services')
        set_of_services, _ = SetOfServices.objects.get_or_create(order_id=pk, name=name)

        service_ids = [int(service) for service in services]
        service_objects = Service.objects.in_bulk(service_ids)
        service_in_set_objects = [ServiceInSet(set=set_of_services, service=service_objects[int(service)]) for service
                                  in services]
        ServiceInSet.objects.bulk_create(service_in_set_objects)

        return redirect('quality_check', pk)


class DefectiveCheckUpdateView(LockedView, UpdateView):
    model = Product
    form_class = DefectiveCheckForm

    def form_valid(self, form):
        set_id = self.request.POST.get('set')
        set_obj = SetOfServices.objects.get(id=set_id)
        services = set_obj.services.all()
        product = self.object
        employer = CustomUser.objects.get(id=form.cleaned_data['employer_id'])
        order = product.order
        client = order.client
        referral = client.referral
        multiplier = 1 if not product.defective_check else -1

        with transaction.atomic():
            good_quality = product.actual_quantity - form.cleaned_data['defective']
            product.defective = form.cleaned_data['defective'] if not product.defective_check else 0
            product.good_quality = good_quality if not product.defective_check else good_quality + form.cleaned_data[
                'defective']
            product.count = product.good_quality + product.defective
            product.defective_check = not product.defective_check
            product.save()

            order_items = Product.objects.filter(order=order)
            total_good_quality = sum(item.good_quality for item in order_items)
            total_defective = sum(item.defective if item.defective is not None else 0 for item in order_items)

            order.good_quality = total_good_quality
            order.defective = total_defective
            order.count = total_good_quality + total_defective

            employer_order, _ = EmployerOrder.objects.get_or_create(order=order, user=employer)
            employer_product, _ = EmployerProduct.objects.get_or_create(product=product, employer=employer)

            client.product_count = order.count

            for service in services:
                service_obj = Service.objects.get(id=service.service.id)
                if service_obj.acceptance:
                    continue
                order_service, _ = OrderService.objects.get_or_create(order=order, service=service_obj)
                service_order, _ = ServiceOrder.objects.get_or_create(order=order, service=service_obj)
                product_service, _ = ProductService.objects.get_or_create(employer_product=employer_product,
                                                                          service=service_obj)

                new_count = product.good_quality if not service.service.before_defective else product.good_quality + \
                                                                                              product.defective

                if service.service.discount:
                    discount = client.discount.percent if client.discount else 0
                    new_amount = (new_count * service_obj.price / 100) * (100 - discount)
                    if referral:
                        referral_money = ((service_obj.price / 100) * 10) * new_count
                        order.referral_money += int(multiplier * referral_money)
                else:
                    new_amount = new_count * service_obj.price

                new_cost = int(new_count * service_obj.cost_price)

                order.amount += int(multiplier * new_amount)
                order.cost_price += int(multiplier * new_cost)

                service_order.count += int(multiplier * new_count)
                service_order.price += int(multiplier * new_amount)

                order_service.count += int(multiplier * new_count)
                order_service.salary += int(multiplier * new_count * service_obj.price)

                employer.money += int(multiplier * new_cost)
                employer.product_count += int(multiplier * new_count)

                client.money += int(multiplier * new_amount)
                client.profit += int(multiplier * (float(new_amount) - float(new_cost)))

                employer_product.service_count += int(multiplier * new_count)
                product_service.count = int(new_count)

                employer_order.service_count += int(multiplier * new_count)
                employer_order.product_count = int(new_count)
                employer_order.salary += int(multiplier * new_cost)

                order.save()
                order_service.save()
                service_order.save()
                employer.save()
                employer_product.save()
                employer_order.save()
                product_service.save()
                client.save()

        return redirect('quality_check', self.object.order.id)

    def form_invalid(self, form):
        print(form.errors)  # В идеале следует использовать логирование
        return redirect('quality_check', self.object.order.id)

    def get_object(self, queryset=None):
        obj = get_object_or_404(Product, pk=self.kwargs['pk'])
        return obj


class InvoiceGenerationView(LockedView, DetailView):
    model = Order
    template_name = 'stages/invoice_generation.html'

    def get_context_data(self, **kwargs):
        context = super(InvoiceGenerationView, self).get_context_data(order_id=self.object.id)
        context['order'] = self.object
        context['all_services'] = Service.objects.filter(single=True)

        context['services'] = ServiceOrder.objects.filter(order_id=self.object.id)
        order = self.object
        service_orders = order.serviceorder_set.order_by('service__type')
        context['service_orders'] = service_orders
        context['workers'] = CustomUser.objects.filter(user_type='worker')
        total_count = self.total_count()
        context['total_count'] = total_count
        context['consumables'] = Consumables.objects.all()
        context['consumables_in_order'] = OrderConsumables.objects.filter(order=self.object)
        return context

    def total_count(self):
        order_products = Order.objects.get(pk=self.object.id)
        products = order_products.products.all()
        total_count = 0

        for product in products:
            total_count += product.actual_quantity
        return total_count


class ApplyDiscountView(LockedView, View):
    def post(self, request):
        order = Order.objects.get(id=self.request.POST.get('order'))
        percent = int(self.request.POST.get('percent'))
        services = ServiceOrder.objects.filter(order=order)
        consumables = OrderConsumables.objects.filter(order=order)
        new_amount = 0
        for sirvice in services:
            if sirvice.service.discount:
                new_amount += (sirvice.price / 100) * (100 - percent)
            else:
                new_amount += sirvice.price
        for consumable in consumables:
            new_amount += consumable.price
        order.amount = round(new_amount)
        order.discount = percent
        order.save()

        return redirect('invoice_generation', order.id)


class AddConsumables(LockedView, View):
    def post(self, request):
        order = self.request.POST.get('order')
        order_obj = Order.objects.get(id=order)
        client = order_obj.client
        consumable = Consumables.objects.get(id=self.request.POST.get('consumable'))
        order_consumable = OrderConsumables.objects.get_or_create(order_id=order, consumable_id=consumable.id)[0]
        old_count = order_consumable.count

        count = int(self.request.POST.get('count'))
        cons_count = consumable.count - count
        if cons_count < 0:
            messages.error(self.request, f"{consumable.name} не достаточно на складе")
            return redirect("invoice_generation", order)
        else:
            consumable.count += old_count
            consumable.count -= count
            consumable.save()
            price = consumable.price * count
            cost_price = consumable.cost_price * count
            order_consumable.count = count
            order_consumable.price = price
            order_consumable.cost_price = cost_price
            order_consumable.save()
            client.money += price
            client.profit += price - cost_price
            client.product_count += count
            client.save()
            order_obj.amount += price
            order_obj.cost_price += cost_price
            order_obj.save()
            return redirect('invoice_generation', order)


class InvoiceGenerationViewGenerate(LockedView, View):

    def convert_som_to_rub(self, amount):
        API_URL = "https://open.er-api.com/v6/latest/KGS"
        response = requests.get(API_URL)

        if response.status_code == 200:
            data = response.json()
            rub_rate = data['rates']['RUB']
            return amount * rub_rate
        else:
            raise Exception("API request failed")

    def post(self, request, order_id):
        file_name = 'blank.xlsx'
        root_directory = 'excel'
        file_path = None

        for root, dirs, files in os.walk(root_directory):
            if file_name in files:
                file_path = os.path.join(root, file_name)
                break

        if not file_path:
            return HttpResponse('File not found')

        wb = load_workbook(file_path)

        sheet = wb.active
        order_obj = Order.objects.get(id=order_id)
        order = order_obj.id
        client = order_obj.client
        date_ = order_obj.date
        name = order_obj.name
        count = order_obj.count
        defective = order_obj.defective
        good_q = order_obj.good_quality

        sheet['B2'] = f'№{order}'
        sheet['B3'] = f'{client}'
        sheet['B4'] = 'ОсОО "Инторг Азия"'
        sheet['B5'] = f'{date_}'
        sheet['B6'] = f'{name}'
        sheet['B7'] = f'{count}'
        sheet['B8'] = f'{defective}'
        sheet['B9'] = f'{good_q}'

        row = 13

        service_orders = ServiceOrder.objects.filter(order=order_id).order_by('service__type__type', 'service__name')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        fill = PatternFill(start_color='000000', end_color='000000', fill_type="solid")

        alignment = Alignment(horizontal='center')

        fill_2 = PatternFill(start_color='D0E0E3', end_color='D0E0E3', fill_type="solid")
        font = Font(color="FFFFFF")
        font_bold = Font(color="FFFFFF", bold=True)

        font_2 = Font(color="000000")
        font_bold_2 = Font(color="000000", bold=True)

        for type_name, group in groupby(service_orders, key=lambda x: x.service.type.type):

            sheet[f'A{row}'].font = font
            sheet[f'A{row}'].border = thin_border
            sheet[f'A{row}'].fill = fill
            sheet[f'B{row}'].fill = fill
            sheet[f'C{row}'].fill = fill
            sheet[f'D{row}'].fill = fill
            sheet[f'A{row}'] = f'{type_name}'
            row += 1

            for service_order in group:
                sheet[f'A{row}'].font = font_2
                sheet[f'A{row}'].border = thin_border
                sheet[f'A{row}'].fill = fill_2
                sheet[f'A{row}'] = f'{service_order.service.name}'
                sheet[f'B{row}'].fill = fill_2
                sheet[f'B{row}'].border = thin_border
                sheet[f'B{row}'] = f'{service_order.count}'

                sheet[f'C{row}'].fill = fill_2
                sheet[f'C{row}'].border = thin_border
                sheet[f'C{row}'].alignment = alignment
                sheet[f'C{row}'] = f'{service_order.service.price}'

                total = service_order.count * service_order.service.price
                sheet[f'D{row}'].fill = fill_2
                sheet[f'D{row}'].alignment = alignment
                sheet[f'D{row}'].border = thin_border
                sheet[f'D{row}'] = f'{total}'
                row += 1

        consumables = OrderConsumables.objects.filter(order_id=order_id)

        if consumables:
            sheet[f'A{row}'].font = font
            sheet[f'A{row}'].border = thin_border
            sheet[f'A{row}'].alignment = alignment
            sheet[f'A{row}'].fill = fill
            sheet[f'B{row}'].fill = fill
            sheet[f'C{row}'].fill = fill
            sheet[f'D{row}'].fill = fill

            sheet[f'A{row}'] = f'Расходные материалы'
            row += 1

            for consumable in consumables:
                sheet[f'A{row}'].font = font_2
                sheet[f'A{row}'].border = thin_border
                sheet[f'A{row}'].fill = fill_2
                sheet[f'A{row}'] = f'{consumable.consumable.name}'
                sheet[f'B{row}'].fill = fill_2
                sheet[f'B{row}'].border = thin_border
                sheet[f'B{row}'] = f'{consumable.count}'

                sheet[f'C{row}'].fill = fill_2
                sheet[f'C{row}'].border = thin_border
                sheet[f'C{row}'].alignment = alignment
                sheet[f'C{row}'] = f'{consumable.consumable.price}'

                total = consumable.count * consumable.consumable.price
                sheet[f'D{row}'].fill = fill_2
                sheet[f'D{row}'].alignment = alignment
                sheet[f'D{row}'].border = thin_border
                sheet[f'D{row}'] = f'{total}'
                row += 1

        sheet[f'A{row}'].font = font_bold
        sheet[f'A{row}'].border = thin_border
        sheet[f'A{row}'].alignment = alignment
        sheet[f'A{row}'].fill = fill
        sheet[f'B{row}'].fill = fill
        sheet[f'C{row}'].fill = fill_2
        sheet[f'C{row}'].font = font_bold_2
        sheet[f'D{row}'].fill = fill_2
        sheet[f'D{row}'].font = font_bold_2
        sheet[f'A{row}'] = f'Способ оплаты'
        sheet[f'C{row}'] = f'скидка {order_obj.discount}%'
        sheet[f'D{row}'] = 'ИТОГО:'
        row += 1
        sheet[f'A{row}'].font = font_bold_2
        sheet[f'A{row}'].border = thin_border
        sheet[f'A{row}'].fill = fill_2
        sheet[f'B{row}'].fill = fill_2
        sheet[f'B{row}'].border = thin_border
        sheet[f'C{row}'].fill = fill_2
        sheet[f'C{row}'].border = thin_border
        sheet[f'D{row}'].font = font_bold_2
        sheet[f'D{row}'].fill = fill_2
        sheet[f'D{row}'].border = thin_border

        sheet[f'A{row}'] = f'Компаньон банк (г.Бишкек) по номеру телефона: +996-507-150-957 П. Жанна Николаевна'
        sheet[f'C{row}'] = ''
        if order_obj.amount == 0:
            sheet[
                f'D{row}'] = f'{order_obj.amount_paid} сом ({round(self.convert_som_to_rub(order_obj.amount_paid), 2)}руб)'
        else:
            sheet[f'D{row}'] = f'{order_obj.amount} сом ({self.convert_som_to_rub(order_obj.amount)}руб)'
        row += 1
        sheet[f'A{row}'].font = font_2
        sheet[f'A{row}'].border = thin_border
        sheet[f'A{row}'].fill = fill_2
        sheet[f'A{row}'] = 'Карта Сбер: 2202-2036-6274-4590 (номер телефона +7-961-226-32-25) К. Алла Ивановна'
        row += 1
        sheet[f'A{row}'].font = font_2
        sheet[f'A{row}'].border = thin_border
        sheet[f'A{row}'].fill = fill_2
        sheet[f'A{row}'] = 'Золотая корона: Камалетдинова Алла Ивановна, КР, г. Бишкек (+996-500-920-908)'

        # Save the workbook
        new_file_path = f'excel/blank{order_id}.xlsx'
        wb.save(new_file_path)
        wb.close()

        # Delete the newly generated file
        max_retries = 3
        retry_delay = 1  # Delay in seconds
        retry_count = 0

        with open(new_file_path, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f'Счет№{order_id} {order_obj.day}/{order_obj.month}/{order_obj.year} {order_obj.client} {order_obj.amount}.xlsx'
            quoted_filename = quote(filename, encoding='utf-8')

            response['Content-Disposition'] = f'attachment; filename="{quoted_filename}"'
        while retry_count < max_retries:
            try:
                os.remove(new_file_path)
                break
            except PermissionError:
                retry_count += 1
                time.sleep(retry_delay)
        return response


class DownloadCargoDetail(View):
    def post(self, request, order_id):
        file_name = 'download_cargo_detail.xlsx'
        root_directory = 'excel'
        file_path = None

        for root, dirs, files in os.walk(root_directory):
            if file_name in files:
                file_path = os.path.join(root, file_name)
                break

        if not file_path:
            return HttpResponse('File not found')

        wb = load_workbook(file_path)

        sheet = wb.active
        order_obj = Order.objects.get(id=order_id)

        row = 3

        products = Product.objects.filter(order_id=order_id)

        total_count = 0
        for product in products:
            sheet[f'A{row}'] = f'{product.name}'
            sheet[f'B{row}'] = f'{product.barcode}'
            sheet[f'C{row}'] = f'{product.article}'
            sheet[f'D{row}'] = f'{product.color}'
            sheet[f'E{row}'] = f'{product.size}'
            sheet[f'F{row}'] = f'{product.good_quality}'
            total_count += product.good_quality
            row += 1
        sheet[f'E{row}'] = 'Итого:'
        sheet[f'F{row}'] = f'{total_count}шт'

        # Save the workbook
        new_file_path = f'excel/blank{order_id}.xlsx'
        wb.save(new_file_path)
        wb.close()

        # Delete the newly generated file
        max_retries = 3
        retry_delay = 1  # Delay in seconds
        retry_count = 0

        with open(new_file_path, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f'Детализация Заказа№{order_obj.id} {order_obj.client}.xlsx'
            quoted_filename = quote(filename, encoding='utf-8')

            response['Content-Disposition'] = f'attachment; filename="{quoted_filename}"'
        while retry_count < max_retries:
            try:
                os.remove(new_file_path)
                break
            except PermissionError:
                retry_count += 1
                time.sleep(retry_delay)
        return response


class DispatchView(LockedView, DetailView):
    model = Order
    template_name = 'stages/dispatch.html'

    def get_context_data(self, **kwargs):
        context = super(DispatchView, self).get_context_data()
        context['products'] = Product.objects.filter(order=self.object)
        context['order_id'] = self.object.id

        return context


class DecreaseProductCountView(LockedView, UpdateView):
    model = Order
    form_class = BarcodeForm

    def form_valid(self, form):
        order = self.object
        barcode = form.cleaned_data['barcode']
        try:
            product = Product.objects.get(barcode=barcode, order=order)
            if product.count > 0:
                product.count -= 1
                product.save()
                return redirect('dispatch', order.id)
            else:
                messages.error(self.request, 'Данного продукта уже нет складе!')
                return redirect('dispatch', order.id)
        except:
            messages.error(self.request, 'Данного продукта нет складе!')
            return redirect('dispatch', order.id)

    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Неверный баркод!')


class AcceptanceNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        if order.stage == 'acceptance':
            order.transition_to_next_stage()

        return redirect(reverse('dashboard'))


class DatabaseLoadingNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        if order.products.count() > 0:
            if order.stage == 'database_loading':
                order.transition_to_next_stage()
            return redirect(reverse('dashboard'))

        else:
            messages.error(self.request, 'Не добавлен ни один продукт!')
            return redirect('import_excel', order.id)


class UnpackingNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        products = order.products.all()
        all_confirmed = all(product.confirmation for product in products)

        if all_confirmed:
            order_items = Product.objects.filter(order=order)
            count = sum(item.actual_quantity for item in order_items)
            order.count = count
            acceptance = Service.objects.get(acceptance=True)
            price = acceptance.price * count
            cost_price = acceptance.cost_price * count
            service_order, _ = ServiceOrder.objects.get_or_create(order=order, service=acceptance)
            service_order.count = count
            service_order.price = price
            service_order.save()
            amount = 0
            client = order.client
            client.money = amount
            client.product_count = count
            client.profit += float(amount) - float(cost_price)
            client.save()
            if acceptance.discount:
                amount += (price / 100) * (100 - order.discount)
                if client.referral:
                    order.referral_money += (price / 100) * 10
            else:
                amount += price

            order.amount += amount
            order.cost_price += cost_price
            order.save()

            employer = CustomUser.objects.get(id=self.request.user.id)
            employer.money += cost_price
            employer.product_count += count
            employer.save()
            order_servise, _ = OrderService.objects.get_or_create(order=order, service=acceptance)
            order_servise.count += count
            order_servise.salary += amount
            order_servise.save()
            employer_order, _ = EmployerOrder.objects.get_or_create(order=order, user_id=self.request.user.id)
            employer_order.service_count += count
            employer_order.product_count += count
            employer_order.salary += cost_price
            employer_order.save()
            for product in products:
                employer_product, _ = EmployerProduct.objects.get_or_create(product=product, employer=employer,
                                                                            product_count=count, service_count=count)
            if order.stage == 'unpacking':
                order.transition_to_next_stage()
                return redirect(reverse('dashboard'))

        else:
            messages.error(self.request, "Не все продукты в заказе подтверждены.")

            return redirect("unpacking", order.id)



class QualityCheckNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        products = order.products.all()

        all_in_work = all(product.in_work for product in products)

        if all_in_work:
            all_defective_checked = all(product.defective_check for product in products)
            if all_defective_checked:
                if order.stage == 'quality_check':
                    order.transition_to_next_stage()
                return redirect(reverse('dashboard'))
            else:
                messages.error(self.request, "Не все проверены на брак.")

                return redirect("quality_check", order.id)
        else:
            messages.error(self.request, "Не все продукты переданы работнкам.")

            return redirect("quality_check", order.id)


class InvoiceGenerationNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        if order.stage == 'invoice_generation':
            order.transition_to_next_stage()
        return redirect(reverse('dashboard'))


class DispatchNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        if order.stage == 'dispatch':
            order.transition_to_next_stage()
        return redirect(reverse('dashboard'))


class DispatchedNextStage(LockedView, View):
    def post(self, request, order_id):
        order = Order.objects.get(id=order_id)
        if order.stage == 'dispatched':
            order.transition_to_next_stage()
        return redirect(reverse('dashboard'))
